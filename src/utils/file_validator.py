import os
import re
from typing import Optional, Union
import openpyxl

class FileValidator:
    """
    Utility class for validating file contents (Text, Excel).
    """

    @staticmethod
    def validate_text_file(path: str, expected_content: str, mode: str = 'exact', encoding: str = 'utf-8') -> None:
        """
        Validates the content of a text file.

        Args:
            path (str): Path to the text file.
            expected_content (str): The expected content or pattern.
            mode (str): Validation mode ('exact', 'contains', 'regex'). Defaults to 'exact'.
            encoding (str): File encoding. Defaults to 'utf-8'.
        
        Raises:
            FileNotFoundError: If the file does not exist.
            AssertionError: If the validation fails.
            ValueError: If an unknown mode is provided.
        """
        if not os.path.exists(path):
            raise FileNotFoundError(f"File not found: {path}")
        
        with open(path, 'r', encoding=encoding) as f:
            content = f.read()
            
        if mode == 'exact':
            assert content == expected_content, f"Content mismatch.\nExpected: '{expected_content}'\nActual:   '{content}'"
        elif mode == 'contains':
            assert expected_content in content, f"Content '{expected_content}' not found in file."
        elif mode == 'regex':
            pattern = re.compile(expected_content)
            assert pattern.search(content), f"Pattern '{expected_content}' not found in file content."
        else:
            raise ValueError(f"Unknown validation mode: {mode}")

    @staticmethod
    def validate_excel_file(path: str, cell: str, expected_value: Union[str, int, float], sheet_name: Optional[str] = None) -> None:
        """
        Validates the value of a specific cell in an Excel file.

        Args:
            path (str): Path to the Excel file.
            cell (str): Cell address (e.g., 'A1').
            expected_value (Union[str, int, float]): The expected value.
            sheet_name (Optional[str]): The name of the sheet. If None, uses the active sheet.

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError: If the sheet is not found.
            AssertionError: If the validation fails.
        """
        if not os.path.exists(path):
             raise FileNotFoundError(f"File not found: {path}")
        
        try:
            workbook = openpyxl.load_workbook(path, data_only=True)
            
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found in workbook. Available sheets: {workbook.sheetnames}")
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
                
            actual_value = sheet[cell].value
            
            # Simple string comparison to handle type differences (e.g. number vs string representation) generally
            # But strict enough for text equality.
            assert str(actual_value) == str(expected_value), f"Cell {cell} mismatch. Expected '{expected_value}' (type: {type(expected_value).__name__}), got '{actual_value}' (type: {type(actual_value).__name__})"
            
        except Exception as e:
            if isinstance(e, (FileNotFoundError, ValueError, AssertionError)):
                raise e
            raise Exception(f"Failed to validate Excel file: {e}")
        finally:
            if 'workbook' in locals():
                workbook.close()
